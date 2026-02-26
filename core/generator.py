import pdfkit
from jinja2 import Environment, FileSystemLoader
import os
from datetime import datetime

# Note: pdfkit requires wkhtmltopdf to be installed on the system.

def generate_final_pdf(marked_up_tables, config):
    """
    Generates a PDF quotation using HTML templates and the marked-up data.
    """
    
    # Process newlines in addresses for HTML
    sender_address_html = config.get('sender_address', '').replace('\n', '<br>')
    recipient_address_html = config.get('recipient_address', '').replace('\n', '<br>')

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
            
            body {{
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
                color: #1e293b;
                background-color: #f8fafc;
                margin: 0;
                padding: 40px;
                line-height: 1.5;
            }}
            .invoice-box {{
                max-width: 900px;
                margin: auto;
                padding: 40px;
                background: #ffffff;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
                border-radius: 12px;
            }}
            .header-table {{
                width: 100%;
                margin-bottom: 40px;
            }}
            .header-table td {{
                vertical-align: top;
            }}
            .title-area h1 {{
                font-size: 42px;
                font-weight: 700;
                color: #0f172a;
                margin: 0;
                letter-spacing: -1px;
            }}
            .title-area p {{
                color: #64748b;
                margin-top: 5px;
                font-size: 14px;
                text-transform: uppercase;
                letter-spacing: 1px;
            }}
            .details-table {{
                width: 100%;
                margin-bottom: 40px;
                border-collapse: collapse;
            }}
            .details-table td {{
                vertical-align: top;
                width: 50%;
                padding: 0;
            }}
            .info-block {{
                background-color: #f1f5f9;
                padding: 24px;
                border-radius: 8px;
            }}
            .info-block.sender {{
                margin-right: 12px;
                border-top: 4px solid #3b82f6; /* Blue accent */
            }}
            .info-block.recipient {{
                margin-left: 12px;
                border-top: 4px solid #10b981; /* Green accent */
            }}
            .info-block h3 {{
                color: #64748b;
                font-size: 12px;
                text-transform: uppercase;
                letter-spacing: 1px;
                margin: 0 0 12px 0;
            }}
            .info-name {{
                font-size: 18px;
                font-weight: 600;
                color: #0f172a;
                margin: 0 0 4px 0;
            }}
            .info-text {{
                color: #475569;
                font-size: 14px;
                margin: 0;
                line-height: 1.6;
            }}
            .job-description-box {{
                background-color: #fef2f2;
                border: 1px solid #fee2e2;
                border-left: 4px solid #ef4444; /* Red accent */
                padding: 20px;
                border-radius: 8px;
                margin-bottom: 40px;
            }}
            .job-description-box h3 {{
                color: #991b1b;
                font-size: 12px;
                text-transform: uppercase;
                letter-spacing: 1px;
                margin: 0 0 8px 0;
            }}
            .job-description-box p {{
                color: #7f1d1d;
                font-size: 14px;
                margin: 0;
                white-space: pre-wrap;
            }}
            .items-section {{
                margin-bottom: 40px;
            }}
            .items-section h2 {{
                font-size: 18px;
                color: #0f172a;
                margin-bottom: 16px;
                padding-bottom: 8px;
                border-bottom: 2px solid #e2e8f0;
            }}
            .dataframe {{
                width: 100%;
                border-collapse: separate;
                border-spacing: 0;
                margin-bottom: 30px;
                font-size: 14px;
            }}
            .dataframe th {{
                background-color: #f8fafc;
                color: #475569;
                font-weight: 600;
                text-align: left;
                padding: 14px 16px;
                border-bottom: 2px solid #cbd5e1;
                border-top: 1px solid #e2e8f0;
            }}
            .dataframe th:first-child {{ border-left: 1px solid #e2e8f0; border-top-left-radius: 6px; }}
            .dataframe th:last-child {{ border-right: 1px solid #e2e8f0; border-top-right-radius: 6px; }}
            
            .dataframe td {{
                padding: 14px 16px;
                color: #334155;
                border-bottom: 1px solid #e2e8f0;
            }}
            .dataframe td:first-child {{ border-left: 1px solid #e2e8f0; }}
            .dataframe td:last-child {{ border-right: 1px solid #e2e8f0; }}
            
            .dataframe tr:last-child td:first-child {{ border-bottom-left-radius: 6px; }}
            .dataframe tr:last-child td:last-child {{ border-bottom-right-radius: 6px; }}
            
            .dataframe tr:nth-child(even) td {{
                background-color: #f8fafc;
            }}
            
            /* Highlight the Grand Total Row */
            .dataframe tr:last-child td {{
                background-color: #f1f5f9;
                font-weight: 700;
                color: #0f172a;
                font-size: 15px;
                border-top: 2px solid #cbd5e1;
            }}
            
            /* Align numbers to the right for cleaner look */
            .dataframe td:not(:first-child), .dataframe th:not(:first-child) {{
                text-align: right;
            }}

            .footer {{
                margin-top: 40px;
                text-align: center;
                color: #94a3b8;
                font-size: 13px;
                padding-top: 24px;
                border-top: 1px solid #e2e8f0;
            }}
        </style>
    </head>
    <body>
        <div class="invoice-box">
            <table class="header-table">
                <tr>
                    <td class="title-area">
                        <h1>QUOTATION</h1>
                        <p>Generated: {datetime.now().strftime("%B %d, %Y")}</p>
                    </td>
                </tr>
            </table>
            
            <table class="details-table">
                <tr>
                    <td>
                        <div class="info-block sender">
                            <h3>From</h3>
                            <p class="info-name">{config.get('sender_name', 'Default Sender')}</p>
                            <p class="info-text">
                                {config.get('sender_phone', '')}{'<br>' if config.get('sender_phone') else ''}
                                {config.get('sender_email', '')}<br>
                                {sender_address_html}
                            </p>
                        </div>
                    </td>
                    <td>
                        <div class="info-block recipient">
                            <h3>Quotation For</h3>
                            <p class="info-name">{config.get('recipient_name', 'Client')}</p>
                            <p class="info-text">
                                {config.get('recipient_contact', '')}<br>
                                {recipient_address_html}
                            </p>
                        </div>
                    </td>
                </tr>
            </table>
            
            {f'''
            <div class="job-description-box">
                <h3>Job Description / Notes</h3>
                <p>{config.get('job_description', '')}</p>
            </div>
            ''' if config.get('job_description') else ''}
            
            <div class="items-section">
                <h2>Line Items</h2>
    """
    
    for df in marked_up_tables:
        # Convert internal tables to HTML, applying the 'dataframe' class defined in CSS
        html_content += df.to_html(index=False, classes='dataframe')
        
    html_content += """
            </div>
            
            <div class="footer">
                <p>Thank you for your business. This quotation is valid for 14 days. Please contact us with any questions.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    try:
        # Generate the PDF from HTML string
        pdf_bytes = pdfkit.from_string(html_content, False)
        return pdf_bytes
    except Exception as e:
        print(f"Error generating PDF (ensure wkhtmltopdf is installed): {e}")
        return None

def generate_excel_from_pdf(tables, config, markup_percentage):
    """
    Generates a professional Excel file from PDF extracted tables, 
    injecting live Excel formulas for the markup calculation.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marked Up Quotation"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_font = Font(size=16, bold=True)
    bold_font = Font(bold=True)
    
    # Add Company Details
    ws.append(["QUOTATION"])
    ws.cell(row=1, column=1).font = title_font
    
    current_date = datetime.now().strftime("%B %d, %Y")
    ws.append([f"Generated: {current_date}"])
    ws.cell(row=2, column=1).font = Font(italic=True, color="64748B")
    
    # Sender details
    ws.append([])
    ws.append(["From:", config.get('sender_name', '')])
    ws.cell(row=4, column=1).font = bold_font
    current_row = 5
    
    if config.get('sender_phone'):
        ws.append(["Phone:", config.get('sender_phone', '')])
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        
    if config.get('sender_email'):
        ws.append(["Email:", config.get('sender_email', '')])
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        
    if config.get('sender_address'):
        ws.append(["Address:", ""])
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        for line in config.get('sender_address').split('\n'):
            if line.strip():
                ws.append(["", line])
                current_row += 1
                
    ws.append([])
    current_row += 1
    
    # Recipient details
    ws.append(["To:", config.get('recipient_name', '')])
    ws.cell(row=current_row, column=1).font = bold_font
    current_row += 1
    
    if config.get('recipient_contact'):
        ws.append(["Contact:", config.get('recipient_contact', '')])
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        
    if config.get('recipient_address'):
        ws.append(["Address:", ""])
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        for line in config.get('recipient_address').split('\n'):
            if line.strip():
                ws.append(["", line])
                current_row += 1
                
    ws.append([])
    current_row += 1
    
    if config.get('job_description'):
        ws.append(["Job Description/Notes:"])
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        
        # Split by newlines so it can print nicely, or just dump into a single cell
        for line in str(config.get('job_description')).splitlines():
            ws.append([line])
            current_row += 1
            
        ws.append([])
        current_row += 1
    
    multiplier = 1 + (markup_percentage / 100)
    
    # Thin Border
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    sum_ranges = []
    last_markup_col_idx = 1
    
    for df in tables:
        # We need the original columns 
        # Add an explicit "Marked Up Total" column
        columns = list(df.columns)
        has_added_total = False
        if "Marked Up Total" not in columns:
            columns.append("Marked Up Total")
            has_added_total = True
            
        # Write Headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        current_row += 1
        start_data_row = current_row
        
        # Write Data
        for _, row in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, 1):
                val = str(row[col_name]).strip()
                
                # Try to clean and convert every cell to float if it looks like a number.
                # This ensures Excel formulas work regardless of which column has numbers.
                if val.lower() in ('nan', 'none', ''):
                    cell = ws.cell(row=current_row, column=col_idx, value="")
                else:
                    try:
                        # Strip currency, commas, spaces
                        clean_str = val.replace('$', '').replace(',', '').replace('€', '').replace('£', '').strip()
                        clean_val = float(clean_str)
                        cell = ws.cell(row=current_row, column=col_idx, value=clean_val)
                        # Optionally apply number formatting only if it's the last original column to avoid formatting quantities as currency
                        if col_idx == len(df.columns):
                            cell.number_format = '#,##0.00'
                    except ValueError:
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        
                cell.border = thin_border
                
            if has_added_total:
                # Add the Excel Formula representing: Original Column * Multiplier
                # We point to the last original column (len(df.columns))
                orig_col_idx = len(df.columns)
                orig_col_letter = openpyxl.utils.get_column_letter(orig_col_idx)
                new_col_letter = openpyxl.utils.get_column_letter(orig_col_idx + 1)
                
                # We need to make sure the targeted cell is actually numeric, otherwise `=C2*1.1` will be #VALUE!
                # We check the cell value that was just written. If it's a string, we just output blank or 0 to be safe.
                target_cell_value = ws.cell(row=current_row, column=orig_col_idx).value
                if isinstance(target_cell_value, (int, float)):
                    formula = f"={orig_col_letter}{current_row}*{multiplier}"
                else:
                    formula = 0 # Fallback if someone put pure text in the price column
                    
                cell = ws.cell(row=current_row, column=orig_col_idx + 1, value=formula)
                cell.number_format = '#,##0.00'
                cell.font = bold_font
                cell.border = thin_border
                
            current_row += 1
            
        end_data_row = current_row - 1
        
        if has_added_total and start_data_row <= end_data_row:
            markup_col_letter = openpyxl.utils.get_column_letter(len(df.columns) + 1)
            sum_ranges.append(f"{markup_col_letter}{start_data_row}:{markup_col_letter}{end_data_row}")
            last_markup_col_idx = len(df.columns) + 1
            
        current_row += 2 # Space between tables
        
    # Generate the Totals Rows
    if sum_ranges:
        subtotal_row_idx = current_row
        ws.cell(row=current_row, column=last_markup_col_idx - 1, value="SUBTOTAL").font = bold_font
        subtotal_formula = f"=SUM({','.join(sum_ranges)})"
        cell = ws.cell(row=current_row, column=last_markup_col_idx, value=subtotal_formula)
        cell.number_format = '#,##0.00'
        cell.font = bold_font
        cell.border = thin_border
        current_row += 1
        
        discount_flat = config.get('discount_flat', 0.0)
        col_letter = openpyxl.utils.get_column_letter(last_markup_col_idx)
        
        # Keep track of which rows to add/subtract for Grand Total
        gt_components = [f"{col_letter}{subtotal_row_idx}"]
        
        if discount_flat > 0:
            disc_row_idx = current_row
            ws.cell(row=current_row, column=last_markup_col_idx - 1, value="DISCOUNT").font = bold_font
            # We insert it as a negative number literal
            cell = ws.cell(row=current_row, column=last_markup_col_idx, value=-discount_flat)
            cell.number_format = '#,##0.00'
            cell.font = Font(color="FF0000", bold=True)
            cell.border = thin_border
            current_row += 1
            # Since the discount cell is negative, we still ADD it to the grand total formula
            gt_components.append(f"{col_letter}{disc_row_idx}")
            
            # The calculation base for tax is (Subtotal + Discount) where Discount is negative
            tax_base_val = f"MAX(0, {col_letter}{subtotal_row_idx}+{col_letter}{disc_row_idx})"
        else:
            tax_base_val = f"{col_letter}{subtotal_row_idx}"
            
        tax_type = config.get('tax_type', 'percentage')
        sales_tax_pct = config.get('sales_tax_percentage', 0.0)
        sales_tax_flat = config.get('sales_tax_flat', 0.0)
        
        if tax_type == 'percentage' and sales_tax_pct > 0:
            tax_row_idx = current_row
            ws.cell(row=current_row, column=last_markup_col_idx - 1, value=f"SALES TAX ({sales_tax_pct}%)").font = bold_font
            tax_formula = f"={tax_base_val}*({sales_tax_pct}/100)"
            cell = ws.cell(row=current_row, column=last_markup_col_idx, value=tax_formula)
            cell.number_format = '#,##0.00'
            cell.font = bold_font
            cell.border = thin_border
            current_row += 1
            gt_components.append(f"{col_letter}{tax_row_idx}")
            
        elif tax_type == 'flat' and sales_tax_flat > 0:
            tax_row_idx = current_row
            ws.cell(row=current_row, column=last_markup_col_idx - 1, value="SALES TAX (Flat)").font = bold_font
            cell = ws.cell(row=current_row, column=last_markup_col_idx, value=sales_tax_flat)
            cell.number_format = '#,##0.00'
            cell.font = bold_font
            cell.border = thin_border
            current_row += 1
            gt_components.append(f"{col_letter}{tax_row_idx}")
            
        ws.cell(row=current_row, column=last_markup_col_idx - 1, value="GRAND TOTAL").font = bold_font
        
        gt_formula = f"={' + '.join(gt_components)}"
            
        cell = ws.cell(row=current_row, column=last_markup_col_idx, value=gt_formula)
        cell.number_format = '#,##0.00'
        cell.font = bold_font
        cell.border = thin_border
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        current_row += 2
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
